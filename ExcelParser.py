import openpyxl
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
import requests
import re
import time
import tkinter as tk
from tkinter import filedialog

def start_script():
    # Load the Excel file
    file_path = input_file_var.get()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Define a fill style for the "ok" cell
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Start the timer
    start_time = time.time()

    # Calculate the total number of rows to process
    total_rows = sheet.max_row - 1  # Subtract 1 for the header row

    # Initialize the counter
    progress_counter = 0

    # Iterate through the rows in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        cell = row[0]  # Get the cell in column E for each row
        if isinstance(cell.value, str) and cell.value.lower().startswith("https"):
            sheet.cell(row=cell.row, column=8, value="ok")  # Write "ok" in column H
            sheet.cell(row=cell.row, column=8).fill = fill  # Apply the fill style

            # Get the URL from the cell
            url = cell.value

            # Scrape the title of the web page using Beautiful Soup
            response = requests.get(url)
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, "html.parser")
                title_tag = soup.find("title")
                if title_tag:
                    title = title_tag.get_text()

                    # Keep only the last two words from the title
                    title_words = title.split()
                    if len(title_words) >= 2:
                        new_title = ' '.join(title_words[-5:-3])
                        # Remove spaces from the string
                        new_title_without_space = new_title.replace(" ", "")

                        # Check if the string contains only digits
                        if new_title_without_space.isdigit():
                            price = int(new_title_without_space)
                            
                        else:
                            price = "bez ceny"                     
                        
                    else:
                        new_title = title

                    sheet.cell(row=cell.row, column=9, value=title)  # Write title in column I
                    sheet.cell(row=cell.row, column=10, value=price)  # Write new title in column J

            # Update the progress counter
            progress_counter += 1

            print(f"\rProgress:{progress_counter}%  ", end="")

    # Save the modified Excel file
    output_path = output_file_var.get()
    wb.save(output_path)

    # Calculate and print the duration
    end_time = time.time()
    duration = end_time - start_time
    print("\nScript duration:", duration, "seconds")
    print("Done. Check the modified Excel file")

def browse_input_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    input_file_var.set(file_path)

def browse_output_file():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_file_var.set(file_path)

# Create the main window
root = tk.Tk()
root.title("Web Scraping Script")

# Create and place widgets
input_file_label = tk.Label(root, text="Input File:")
input_file_label.pack()
input_file_var = tk.StringVar(value="INPUT.xlsx")
input_file_entry = tk.Entry(root, textvariable=input_file_var)
input_file_entry.pack()
input_file_button = tk.Button(root, text="Browse", command=browse_input_file)
input_file_button.pack()

output_file_label = tk.Label(root, text="Output File:")
output_file_label.pack()
output_file_var = tk.StringVar(value="EDITED_INPUT.xlsx")
output_file_entry = tk.Entry(root, textvariable=output_file_var)
output_file_entry.pack()
output_file_button = tk.Button(root, text="Browse", command=browse_output_file)
output_file_button.pack()

start_button = tk.Button(root, text="Start", command=start_script)
start_button.pack()

cancel_button = tk.Button(root, text="Cancel", command=root.destroy)
cancel_button.pack()

# Start the GUI event loop
root.mainloop()
